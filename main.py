import datetime
import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk
from dotenv import load_dotenv

load_dotenv()

# Define the correct username and password
correct_username = os.getenv('correct_username')
correct_password = os.getenv('correct_password')


# Function to check login credentials
def check_login(username, password):
    return username == correct_username and password == correct_password

# Function to open the main application
def open_main_app():
    login_window.destroy()  # Close the login window
    initialize_app()

# Function to handle the login button click
def login():
    username = username_entry.get()
    password = password_entry.get()

    if check_login(username, password):
        open_main_app()
    else:
        messagebox.showerror("Login Failed", "Invalid username or password")


# Create the main application window and the rest of your code
def initialize_app():

    #function to extract and search for the needed data ( list of numbers)
    def extract_data(file_path, target_word, stop_word):
        count = 0
        start_index = -1

        # Read the text from the file
        with open(file_path, 'r') as file:
            lines = file.readlines()

        # Go through the lines
        for index, line in enumerate(lines):
            if target_word in line:
                count += 1

                # Check if the word is founded
                if count == 1:
                    start_index = index
                    break

        # Extract data starting from the third occurrence until the stop word is found
        if start_index != -1:
            extracted_data = []

            for i in range(start_index + 1, len(lines)):
                if stop_word in lines[i]:
                    break
                #mahou each ligne feha 1,2,3,4... w feha aka valeurs. so for each ligne ychouf fama les deux valeurs walee so it's kinda of cleaning
                words = lines[i].split()
                if len(words) >= 2:
                    extracted_data.append(words[:2])

            return extracted_data

        return None  # The word was not found

    #function called to clear the frames when an other button is called
    def clear_graph_frame():
        for widget in graph_frame.winfo_children():
            widget.destroy()



    def gab1(file_path):
        count = 0
        # Load the workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Check from B2 (1) to B32(31)
        # go through all the data
        for row in range(2, 33):
            value = float(sheet[f'B{row}'].value)
            if value >= -120:
                count += 1

        if count > 3:
            result1 = "Plus de 3 valeurs sont plus de -120 ----- gab1 KO"
        else:
            result1 = " OK"
        # Calculate the average of B2 to B32
        values = [float(sheet[f'B{i}'].value) for i in range(2, 33)]
        average = sum(values) / len(values)
        # Check the average value
        if average < -130:
            result2 = " OK"
        else:
            result2 = " Moyenne =", str(round(average)), " ----- gab1 KO"

        if "KO" in result1:
            return result1
        elif "KO" in result2:
            return result2
        else:
            return "OK"

    def gab2(file_path):
        count = 0
        # Load the workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Check from B33 (32) to B4096(4095)
        # go through all the data
        for row in range(33, 4097):
            value = float(sheet[f'B{row}'].value)
            if value >= -138:
                count += 1

        if count > 10:
            result1 = "Plus de 10 valeurs sont plus de -138 ----- gab2 KO"
        else:
            result1 = " OK"
        # Calculate the average of B33 to B4096
        values = [float(sheet[f'B{i}'].value) for i in range(33, 4097)]
        average = sum(values) / len(values)
        # Check average value
        if average < -143:
            result2 = " OK"
        else:
            result2 = "Moyenne=", str(round(average)), " ----- gab2 KO"

        if "KO" in result1:
            return result1
        elif "KO" in result2:
            return result2
        else:
            return "OK"

    def gab3(file_path):
        count = 0
        # Load the workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Check from B4097(4096) to B8192(8191)
        for row in range(4097, 8193):
            value = float(sheet[f'B{row}'].value)
            if value >= -132:
                count += 1

        if count > 3:
            result1 = "Plus de 3 valeurs sont plus de -132 ----- gab1 KO"
        else:
            result1 = " OK"
        # Calculate the average of B4097 to B8192
        values = [float(sheet[f'B{i}'].value) for i in range(4097, 8193)]
        average = sum(values) / len(values)
        # Check average value
        if average < -137:
            result2 = " OK"
        else:
            result2 = "Moyenne=", str(round(average)), " ----- gab3 KO"

        if "KO" in result1:
            return result1
        elif "KO" in result2:
            return result2
        else:
            return "OK"

    #fucntion to check just one file
    from openpyxl import load_workbook

    def analyze_file(file_path):
        # Check if the file is a text file
        if not file_path.endswith(".txt"):
            return "Not a valid text file"

        # Extract the file name from the file path
        file_name = os.path.basename(file_path)

        # Define the target and stop words
        target_word = 'number'
        stop_word = 'root@sh31b_cp_l0:~#'

        # Call the extract data function
        extracted_data = extract_data(file_path, target_word, stop_word)

        # If data is found, create an Excel file for analysis
        if extracted_data:
            # Create a new excel
            workbook = Workbook()
            sheet = workbook.active

            # Write the extracted data vertically
            for row_index, row_data in enumerate(extracted_data):
                for col_index, value in enumerate(row_data):
                    sheet.cell(row=row_index + 1, column=col_index + 1, value=value)

            # Save the workbook with the same name as the input .txt file
            excel_file_path = os.path.splitext(file_path)[0] + ".xlsx"
            workbook.save(excel_file_path)

            # Analyze the Excel file
            gab1_result = gab1(excel_file_path)
            gab2_result = gab2(excel_file_path)
            gab3_result = gab3(excel_file_path)

            # Initialize a list to store detailed failure reasons
            failure_reasons = []

            # Initialize a list to store detailed gab results
            gab_results = []

            # Check for gab1 failure
            if "KO" in gab1_result:
                failure_reasons.append("gab1")
                gab_results.append(f"gab1: {gab1_result}")

            # Check for gab2 failure
            if "KO" in gab2_result:
                failure_reasons.append("gab2")
                gab_results.append(f"gab2: {gab2_result}")

            # Check for gab3 failure
            if "KO" in gab3_result:
                failure_reasons.append("gab3")
                gab_results.append(f"gab3: {gab3_result}")

            # Create a detailed message for the failure reasons
            if failure_reasons:
                failure_message = f"{file_name} is KO due to the following reasons:\n\n"
                failure_message += "\n\n".join(gab_results) + "\n\n"
                sheet.cell(row=1, column=col_index + 2, value=failure_message)  # Write the failure message to Excel

                # Save the workbook again with the updated data
                workbook.save(excel_file_path)

                return failure_message
            else:
                sheet.cell(row=1, column=col_index + 2, value="OK")  # Write "OK" to Excel

                # Save the workbook again with the updated data
                workbook.save(excel_file_path)

                return f"{file_name} is OK"
        else:
            return f"{file_name} Data not found"

    #function to check all the files in a folder. this function is called when the choose folder button is clicked
    def analyze_files_in_folder(folder_path):
        #call the clear graphe frame function
        clear_graph_frame()
        #get all the .txt files in that folder
        text_files = glob.glob(os.path.join(folder_path, "*.txt"))
        #get the number of toal files
        total_files = len(text_files)
        #we r gonna use that to calculate later
        ok_files, ko_files = 0, 0

        # Lists to store file names and OK/KO results
        file_names = []
        ok_or_ko = []

        # Lists to store reasons for "KO"
        ko_reasons = []

        for file_path in text_files:
            #get the name of the file
            file_name = os.path.basename(file_path)
            file_names.append(file_name)
            #call the analyze file function
            result = analyze_file(file_path)
            if "KO" in result:
                ok_or_ko.append("KO")
                ko_files += 1
                if "gab1" in result:
                    ko_reasons.append(f"{result}")
                elif "gab2" in result:
                    ko_reasons.append(f"{result}")
                elif "gab3" in result:
                    ko_reasons.append(f"{result}")
            else:
                ok_or_ko.append("OK")
                ok_files += 1

        # Display results inside the Text widget
        total_percentage = (total_files - ko_files) / total_files * 100

        result_text = f"Total Files: {total_files}\n"
        result_text += f"OK Files: {ok_files}, KO Files: {ko_files}\n"
        result_text += f"Total OK Percentage: {total_percentage:.2f}%"

        # Display reasons for "KO" in the text result
        if ko_reasons:
            result_text += "\nReasons for KO:\n"
            result_text += "\n".join(ko_reasons)

        text_results.delete(1.0, tk.END)  # Clear existing text

        # Create results list matching the first function
        results = []
        for file_name, status in zip(file_names, ok_or_ko):
            if status == "OK":
                results.append(f"File: {file_name}, Result: OK\n\n")
            else:
                results.append(f"{ko_reasons.pop(0)}\n\n")

        # Insert results into the Text widget
        if results:
            text_results.insert(tk.END, "\n".join(results))
        else:
            text_results.insert(tk.END, "No files found.")

        # Traçage pour le graphe
        labels = ['OK', 'KO']
        sizes = [ok_files, ko_files]
        explode = (0, 0.1)  # explode KO

        fig, ax = plt.subplots(1, 2, figsize=(10, 6))
        fig.suptitle("File Analysis Results")

        # le graphe e cercle li howa Pie esmou
        ax[0].pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%', shadow=True, startangle=90)
        ax[0].axis('equal')  # ensures that pie is drawn as a circle.
        ax[0].set_title("OK/KO Distribution")

        # Create the bar graphe
        bars = ax[1].bar(range(len(file_names)), [1 if "OK" in ok else 0.1 for ok in ok_or_ko],
                         color=['g' if "OK" in ok else 'r' for ok in ok_or_ko], width=0.8)

        ax[1].set_xlabel("File Names")
        ax[1].set_title("File Status", fontsize=10)

        # Set the font size for the X-axis labels
        ax[1].tick_params(axis='x', labelrotation=45, labelsize=8)

        # Set Y-axis labels
        ax[1].set_yticks([0.1, 1])
        ax[1].set_yticklabels(["KO", "OK"])

        # Add labels below the bars
        ax[1].set_xticks(range(len(file_names)))
        ax[1].set_xticklabels(file_names, rotation=45, fontsize=8, ha='right')

        plt.subplots_adjust(wspace=0.5)

        # Display the plot (Traçage) in the Tkinter window
        canvas = FigureCanvasTkAgg(fig, master=graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

    #function that will open you the file explorer to choose the folder
    def choose_folder():
        folder_path = filedialog.askdirectory()
        if folder_path:
            analyze_files_in_folder(folder_path)

    #the text results for the 1 file
    def analyze_file_button():
        clear_graph_frame()
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file_path:
            result = analyze_file(file_path)
            text_results.delete(1.0, tk.END)  # Clear existing text
            text_results.insert(tk.END, result)


    def analyze_files_in_date_range(folder_path, start_date, end_date):
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            return "Invalid date format. Please use YYYY-MM-DD format."

        text_files = glob.glob(os.path.join(folder_path, "*.txt"))
        results = []

        ok_files, ko_files = 0, 0
        ko_gab1, ko_gab2, ko_gab3 = 0, 0, 0

        # Lists to store file names and OK/KO results
        file_names = []
        ok_or_ko = []

        for file_path in text_files:
            # Get the creation time of the file
            creation_time = datetime.datetime.fromtimestamp(os.path.getctime(file_path))

            # Check if the creation time is within the specified date range
            if start_date <= creation_time <= end_date:
                file_name = os.path.basename(file_path)
                file_names.append(file_name)

                result = analyze_file(file_path)
                if "KO" in result:
                    ok_or_ko.append("KO")
                    ko_files += 1
                    if "gab1" in result:
                        ko_gab1 += 1
                    elif "gab2" in result:
                        ko_gab2 += 1
                    elif "gab3" in result:
                        ko_gab3 += 1
                else:
                    ok_or_ko.append("OK")
                    ok_files += 1

                # Append the result to the results list
                results.append(f"File: {file_name}, Result: {result}")

        # Display results inside the Text widget
        total_percentage = (len(text_files) - ko_files) / len(text_files) * 100
        ko_gab1_percentage = ko_gab1 / ko_files * 100 if ko_files > 0 else 0
        ko_gab2_percentage = ko_gab2 / ko_files * 100 if ko_files > 0 else 0
        ko_gab3_percentage = ko_gab3 / ko_files * 100 if ko_files > 0 else 0

        result_text = f"OK Files: {ok_files}, KO Files: {ko_files}\n"
        result_text += f"Total OK Percentage: {total_percentage:.2f}%\n"
        result_text += f"KO Gab1 Percentage: {ko_gab1_percentage:.2f}%\n"
        result_text += f"KO Gab2 Percentage: {ko_gab2_percentage:.2f}%\n"
        result_text += f"KO Gab3 Percentage: {ko_gab3_percentage:.2f}%"

        text_results.delete(1.0, tk.END)  # Clear existing text
        text_results.insert(tk.END, result_text)

        # Traçage again
        labels = ['OK', 'KO']
        sizes = [ok_files, ko_files]
        explode = (0, 0.1)  # explode KO

        fig, ax = plt.subplots(1, 2, figsize=(10, 6))
        fig.suptitle("File Analysis Results")

        # Pie chart for OK/KO
        ax[0].pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%', shadow=True, startangle=90)
        ax[0].axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        ax[0].set_title("OK/KO Distribution")

        # Create the bar chart
        bars = ax[1].bar(range(len(file_names)), [1 if ok == "OK" else 0.1 for ok in ok_or_ko],
                         color=['g' if ok == "OK" else 'r' for ok in ok_or_ko], width=0.8)  # Adjust the width as needed

        ax[1].set_xlabel("File Names")
        ax[1].set_title("File Status", fontsize=10)  # Adjust the fontsize as needed

        # Set the font size for the X-axis labels
        ax[1].tick_params(axis='x', labelrotation=45, labelsize=8)  # Adjust labelrotation and labelsize as needed

        # Set Y-axis labels
        ax[1].set_yticks([0.1, 1])
        ax[1].set_yticklabels(["KO", "OK"])

        # Add labels below the bars
        ax[1].set_xticks(range(len(file_names)))
        ax[1].set_xticklabels(file_names, rotation=45, fontsize=8, ha='right')

        plt.subplots_adjust(wspace=0.5)

        # Display the plot chouhira Traçage in the Tkinter window
        canvas = FigureCanvasTkAgg(fig, master=graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        if not results:
            return "No files found in the specified date range."

        return "\n".join(results)


    # Function to handle date range selection
    def select_date_range():
        clear_graph_frame()
        start_date = start_date_entry.get()
        end_date = end_date_entry.get()

        if not start_date or not end_date:
            messagebox.showerror("Date Range Error", "Please enter both start and end dates.")
            return

        folder_path = filedialog.askdirectory()
        if folder_path:
            results = analyze_files_in_date_range(folder_path, start_date, end_date)
            text_results.delete(1.0, tk.END)  # Clear existing text
            text_results.insert(tk.END, results)

    #lena efhem wa7dek aka buttons w labels xD
    window = tk.Tk()
    window.title("Sagem Controlling Devices Testing")
    logo_image = tk.PhotoImage(file=r"sagem-removebg-preview.png")
    logo_image = logo_image.subsample(2, 2)
    # Create a label for the logo and display it in the top-left corner
    logo_label = tk.Label(window, image=logo_image)
    logo_label.place(relx=0, rely=0, anchor="nw",x=10,y=10)

    # Make the main window fullscreen
    window.geometry("1920x1080")

    # Create a title label at the top middle
    title_label = tk.Label(window, text="Remote Controlling Devices \nTesting and Analysis",fg="#0982ec", font=("Gabarito", 26))
    title_label.place(relx=0.5, rely=0.01, anchor="n")

    # Create a frame for the buttons on the top right
    button_frame = tk.Frame(window)
    button_frame.place(relx=1, rely=0, anchor="ne",x=-120,y=60)

    # Create the "Choose Folder" button and place it at the top right
    folder_button = tk.Button(button_frame, text="Choose Folder", command=choose_folder, font=("Helvetica", 12))
    folder_button.grid(row=0, column=0, pady=5)

    # Create the "Choose File" button and place it next to the "Choose Folder" button
    analyze_file_button = tk.Button(button_frame, text="Choose File", command=analyze_file_button,
                                    font=("Helvetica", 12))
    analyze_file_button.grid(row=0, column=1, padx=5, pady=5)

    # Create a frame for the date input and button
    date_frame = tk.Frame(window)
    date_frame.place(relx=1, rely=0.1, anchor="ne",x=-70, y=20)

    start_date_label = tk.Label(date_frame, text="Start Date (YYYY-MM-DD):", font=("Helvetica", 10))
    start_date_label.grid(row=0, column=0, sticky="w")

    start_date_entry = tk.Entry(date_frame, font=("Helvetica", 10))
    start_date_entry.grid(row=1, column=0, padx=10, sticky="w")

    end_date_label = tk.Label(date_frame, text="End Date (YYYY-MM-DD):", font=("Helvetica", 10))
    end_date_label.grid(row=2, column=0, sticky="w")

    end_date_entry = tk.Entry(date_frame, font=("Helvetica", 10))
    end_date_entry.grid(row=3, column=0, padx=10, sticky="w")

    date_range_button = tk.Button(date_frame, text="Select Date Range", command=select_date_range,
                                  font=("Helvetica", 10))
    date_range_button.grid(row=0, column=1, rowspan=4, padx=10, pady=10, sticky="w")



    # Create a frame for the graphs (circle and bar) on the left
    graph_container = tk.Frame(window)
    graph_container.place(relx=0.37, rely=0.52, anchor="center")

    # Create a frame for the graphs (circle and bar) inside the container frame
    graph_frame = tk.Frame(graph_container, bd=2, relief="solid")
    graph_frame.pack(padx=0, pady=0)  # Adjust padx and pady as needed

    # Create the frame for the text results on the right
    text_results_frame = tk.Frame(window,bd=2, relief="solid")
    text_results_frame.place(relx=0.85, rely=0.3, anchor="n")

    # Text results
    text_results = tk.Text(text_results_frame, height=30, width=50)
    text_results.pack(side=tk.RIGHT)

    window.mainloop()


# Function to clear the placeholder text when entry field is clicked
def clear_placeholder(event):
    if username_entry.get() == "Username":
        username_entry.delete(0, "end")
        username_entry["style"] = "TEntry"

#and this one for password
def clear_placeholder_password(event):
    if password_entry.get() == "Password":
        password_entry.delete(0, "end")
        password_entry["style"] = "TEntry"
        password_entry.configure(show="*")

# Create the login window
login_window = tk.Tk()
login_window.title("Login")
login_window.geometry("720x315")
icon_path = "sagem.ico"
login_window.iconbitmap(default=icon_path)

# Load the original image
original_image = Image.open(r"Sagemcom.png")

resized_image = original_image.resize((720, 315))

# Convert the resized image to a format suitable for Tkinter
background_image = ImageTk.PhotoImage(resized_image)

# Create a label with the background image and make it cover the whole window
bg_label = tk.Label(login_window, image=background_image)
bg_label.place(relwidth=1, relheight=1)

# Username Entry with Placeholder
username_entry = ttk.Entry(login_window, font=("Helvetica", 10), justify="center", width=20)
username_entry.insert(0, "Username")  # Set the placeholder text
username_entry.bind("<FocusIn>", clear_placeholder)
username_entry.place(relx=0.5, rely=0.4, anchor="center")

# Password Entry with Placeholder
password_entry = ttk.Entry(login_window, font=("Helvetica", 10), justify="center", show="*", width=20)
password_entry.insert(0, "Password")  # Set the placeholder text
password_entry.bind("<FocusIn>", clear_placeholder_password)
password_entry.place(relx=0.5, rely=0.5, anchor="center")

# Login Button with Rounded Corners and Background Color
login_button = ttk.Button(
    login_window, text="Login", command=login)
login_button.place(relx=0.5, rely=0.6, anchor="center")

# Start the login window
login_window.mainloop()
